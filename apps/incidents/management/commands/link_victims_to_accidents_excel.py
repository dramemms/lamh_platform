from openpyxl import load_workbook
from openpyxl.utils.datetime import from_excel
from django.core.management.base import BaseCommand

from apps.incidents.models import Accident
from apps.victims.models import Victim
from apps.geo.models import Region, Cercle, Commune

from datetime import datetime, date
import re
import unicodedata


def clean(value):
    if value is None:
        return ""
    return str(value).replace("\xa0", " ").strip()


def clean_none(value):
    v = clean(value)
    return v if v else None


def normalize(value):
    value = clean(value).lower()
    value = value.replace("_cercle", "").replace("_cerle", "")
    value = unicodedata.normalize("NFKD", value)
    value = "".join(c for c in value if not unicodedata.combining(c))
    value = re.sub(r"\s+", " ", value)
    return value.strip()


def get_value(data, column_name):
    target = normalize(column_name)
    for key, value in data.items():
        if normalize(key) == target:
            return value
    return None


def clean_date(value):
    if value in [None, ""]:
        return None
    if isinstance(value, datetime):
        return value.date()
    if isinstance(value, date):
        return value
    if isinstance(value, (int, float)):
        try:
            return from_excel(value).date()
        except Exception:
            return None

    value = clean(value)

    for fmt in ("%d/%m/%Y", "%Y-%m-%d", "%d-%m-%Y"):
        try:
            return datetime.strptime(value, fmt).date()
        except Exception:
            pass

    return None


def clean_int(value):
    if value in [None, ""]:
        return None

    value = clean(value)
    match = re.search(r"\d+", value)
    if match:
        return int(match.group())

    return None


def get_status_value():
    if "status" not in [f.name for f in Victim._meta.fields]:
        return None

    choices = [c[0] for c in Victim._meta.get_field("status").choices]

    if "SUBMITTED" in choices:
        return "SUBMITTED"
    if "SOUMIS" in choices:
        return "SOUMIS"

    return choices[0] if choices else None


def set_if_exists(obj, field_name, value):
    if field_name in [f.name for f in obj._meta.fields]:
        setattr(obj, field_name, value)


def find_by_normalized_name(model, name, parent_field=None, parent_obj=None):
    name_norm = normalize(name)
    if not name_norm:
        return None

    qs = model.objects.all()

    if parent_field and parent_obj:
        qs = qs.filter(**{parent_field: parent_obj})

    for obj in qs:
        if normalize(getattr(obj, "name", "")) == name_norm:
            return obj

    return None


def find_region(name):
    return find_by_normalized_name(Region, name)


def find_cercle(name, region=None):
    cercle = find_by_normalized_name(Cercle, name, "region", region)
    if cercle:
        return cercle
    return find_by_normalized_name(Cercle, name)


def find_commune(name, cercle=None):
    commune = find_by_normalized_name(Commune, name, "cercle", cercle)
    if commune:
        return commune
    return find_by_normalized_name(Commune, name)


def find_accident_by_context(accident_date, region, cercle, commune, village):
    qs = Accident.objects.all()

    if accident_date:
        qs = qs.filter(accident_date=accident_date)

    if region:
        qs = qs.filter(region=region)

    if cercle:
        qs = qs.filter(cercle=cercle)

    if commune:
        qs = qs.filter(commune=commune)

    exact = list(qs)

    if len(exact) == 1:
        return exact[0], "match_date_geo"

    village_norm = normalize(village)
    if village_norm:
        village_matches = [
            a for a in exact
            if normalize(getattr(a, "village", "")) == village_norm
        ]
        if len(village_matches) == 1:
            return village_matches[0], "match_date_geo_village"

    if len(exact) > 1:
        return exact[0], "match_multiple_first"

    return None, "not_found"


class Command(BaseCommand):
    help = "Importer/lier les victimes aux accidents par contexte"

    def add_arguments(self, parser):
        parser.add_argument("file_path", type=str)

    def handle(self, *args, **options):
        file_path = options["file_path"]

        wb = load_workbook(file_path, data_only=True)

        if "Victim form" in wb.sheetnames:
            ws = wb["Victim form"]
        else:
            ws = wb.active

        self.stdout.write(self.style.WARNING(f"Feuille utilisée : {ws.title}"))

        headers = [clean(cell.value) for cell in ws[1]]

        created = 0
        updated = 0
        not_found = 0
        errors = 0

        for idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
            data = dict(zip(headers, row))

            if not any(row):
                continue

            try:
                victim_id = clean(get_value(data, "1.1. ID de la victime"))

                accident_date = clean_date(get_value(data, "1.4. Date de l'accident"))
                region_name = clean(get_value(data, "4.2. Région"))
                cercle_name = clean(get_value(data, "4.3. Cercle"))
                commune_name = clean(get_value(data, "4.4. Commune"))
                village = clean(get_value(data, "4.5. Village / Quartier"))

                region = find_region(region_name)
                cercle = find_cercle(cercle_name, region)
                commune = find_commune(commune_name, cercle)

                accident, mode = find_accident_by_context(
                    accident_date=accident_date,
                    region=region,
                    cercle=cercle,
                    commune=commune,
                    village=village,
                )

                if not accident:
                    not_found += 1
                    self.stdout.write(
                        self.style.ERROR(
                            f"Ligne {idx} : accident introuvable | date={accident_date} | région={region_name} | cercle={cercle_name} | commune={commune_name} | village={village}"
                        )
                    )
                    continue

                victim = None
                if victim_id:
                    victim = Victim.objects.filter(victim_id__iexact=victim_id).first()

                is_new = False
                if not victim:
                    victim = Victim()
                    is_new = True

                set_if_exists(victim, "victim_id", victim_id)
                set_if_exists(victim, "accident", accident)
                set_if_exists(victim, "accident_reference", accident.reference)

                status_value = get_status_value()
                if status_value:
                    set_if_exists(victim, "status", status_value)

                set_if_exists(victim, "report_date", clean_date(get_value(data, "1.2. Date du rapport")))
                set_if_exists(victim, "accident_date", accident_date)
                set_if_exists(victim, "reported_by", clean_none(get_value(data, "1.5. Raporté par")))
                set_if_exists(victim, "organisation", clean_none(get_value(data, "1.6. Organisation")))
                set_if_exists(victim, "position", clean_none(get_value(data, "1.7. Poste ")))
                set_if_exists(victim, "team", clean_none(get_value(data, "1.8. Equipe ")))

                set_if_exists(victim, "last_name", clean_none(get_value(data, "2.2. Nom de la victime")))
                set_if_exists(victim, "first_name", clean_none(get_value(data, "2.3. Prénom de la victime ")))
                set_if_exists(victim, "victim_type", clean_none(get_value(data, "2.4.  Type de victime")))
                set_if_exists(victim, "father_name", clean_none(get_value(data, "2.5. Nom du père ")))
                set_if_exists(victim, "mother_name", clean_none(get_value(data, "2.6. Nom de la mère")))
                set_if_exists(victim, "nationality", clean_none(get_value(data, "2.7. Nationalité")))
                set_if_exists(victim, "gender", clean_none(get_value(data, "2.13. Sexe")))
                set_if_exists(victim, "age", clean_int(get_value(data, "2.12.3.  Âge")))
                set_if_exists(victim, "contact", clean_none(get_value(data, "2.17. Contact")))

                set_if_exists(victim, "country", clean_none(get_value(data, "4.1. Pays")))
                set_if_exists(victim, "region", region or accident.region)
                set_if_exists(victim, "cercle", cercle or accident.cercle)
                set_if_exists(victim, "commune", commune or accident.commune)
                set_if_exists(victim, "village", village)

                victim.save()

                if is_new:
                    created += 1
                else:
                    updated += 1

                self.stdout.write(
                    self.style.SUCCESS(
                        f"Ligne {idx} OK → victime {victim_id} liée à {accident.reference} ({mode})"
                    )
                )

            except Exception as e:
                errors += 1
                self.stdout.write(
                    self.style.ERROR(f"Ligne {idx} ERREUR → {type(e).__name__}: {e}")
                )

        self.stdout.write(self.style.SUCCESS(f"Victimes créées : {created}"))
        self.stdout.write(self.style.SUCCESS(f"Victimes mises à jour : {updated}"))
        self.stdout.write(self.style.WARNING(f"Accidents introuvables : {not_found}"))
        self.stdout.write(self.style.WARNING(f"Erreurs : {errors}"))