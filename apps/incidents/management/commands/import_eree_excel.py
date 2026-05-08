from openpyxl import load_workbook
from openpyxl.utils.datetime import from_excel
from django.core.management.base import BaseCommand
from django.utils.crypto import get_random_string

from apps.eree.models import EREESession
from apps.geo.models import Region, Cercle, Commune

from datetime import datetime, date
from decimal import Decimal, InvalidOperation
import re
import unicodedata


def clean(value):
    if value is None:
        return ""
    return str(value).replace("\xa0", " ").strip()


def clean_none(value):
    value = clean(value)
    return value if value else None


def normalize(value):
    value = clean(value).lower()
    value = unicodedata.normalize("NFKD", value)
    value = "".join(c for c in value if not unicodedata.combining(c))
    value = re.sub(r"\s+", " ", value)
    return value.strip()


def get_value(data, column_name):
    target = normalize(column_name)
    for key, value in data.items():
        if normalize(key) == target:
            return value
    return None


def clean_date(value):
    if value in [None, ""]:
        return None

    if isinstance(value, datetime):
        return value.date()

    if isinstance(value, date):
        return value

    if isinstance(value, (int, float)):
        try:
            return from_excel(value).date()
        except Exception:
            return None

    value = clean(value)

    for fmt in ("%d/%m/%Y", "%Y-%m-%d", "%d-%m-%Y"):
        try:
            return datetime.strptime(value, fmt).date()
        except Exception:
            pass

    return None


def clean_int(value):
    if value in [None, ""]:
        return 0

    value = clean(value).replace(",", ".")

    try:
        return int(float(value))
    except Exception:
        match = re.search(r"\d+", value)
        return int(match.group()) if match else 0


def clean_decimal(value, coord_type=None):
    """
    Nettoyage sécurisé pour DecimalField latitude/longitude.
    Retourne None si la valeur est invalide.
    """
    if value in [None, ""]:
        return None

    value = clean(value).replace(",", ".")

    if value.lower() in ["nan", "none", "null", "-", "n/a", "na"]:
        return None

    if not re.match(r"^-?\d+(\.\d+)?$", value):
        return None

    try:
        d = Decimal(value).quantize(Decimal("0.000001"))
    except (InvalidOperation, ValueError):
        return None

    if coord_type == "lat" and not (Decimal("-90") <= d <= Decimal("90")):
        return None

    if coord_type == "lng" and not (Decimal("-180") <= d <= Decimal("180")):
        return None

    return d


def generate_unique_eree_reference(session_date=None):
    if session_date:
        date_part = session_date.strftime("%Y%m%d")
    else:
        date_part = datetime.now().strftime("%Y%m%d")

    while True:
        reference = f"EREE-{date_part}-{get_random_string(8).upper()}"
        if not EREESession.objects.filter(reference=reference).exists():
            return reference


def set_if_exists(obj, field_name, value):
    if field_name in [f.name for f in obj._meta.fields]:
        setattr(obj, field_name, value)


def find_by_normalized_name(model, name, parent_field=None, parent_obj=None):
    name_norm = normalize(name)

    if not name_norm:
        return None

    qs = model.objects.all()

    if parent_field and parent_obj:
        qs = qs.filter(**{parent_field: parent_obj})

    for obj in qs:
        if normalize(getattr(obj, "name", "")) == name_norm:
            return obj

    return None


def find_region(name):
    return find_by_normalized_name(Region, name)


def find_cercle(name, region=None):
    cercle = find_by_normalized_name(Cercle, name, "region", region)
    if cercle:
        return cercle
    return find_by_normalized_name(Cercle, name)


def find_commune(name, cercle=None):
    commune = find_by_normalized_name(Commune, name, "cercle", cercle)
    if commune:
        return commune
    return find_by_normalized_name(Commune, name)


def get_submitted_value(field_name):
    fields = [f.name for f in EREESession._meta.fields]

    if field_name not in fields:
        return None

    field = EREESession._meta.get_field(field_name)

    if not field.choices:
        return "SUBMITTED"

    choices = [c[0] for c in field.choices]

    if "SUBMITTED" in choices:
        return "SUBMITTED"

    if "SOUMIS" in choices:
        return "SOUMIS"

    return choices[0] if choices else "SUBMITTED"


def calculate_totals(obj):
    pdi_fields = [
        "pdi_boys_0_5", "pdi_boys_0_5_dis",
        "pdi_girls_0_5", "pdi_girls_0_5_dis",
        "pdi_boys_6_14", "pdi_boys_6_14_dis",
        "pdi_girls_6_14", "pdi_girls_6_14_dis",
        "pdi_boys_15_17", "pdi_boys_15_17_dis",
        "pdi_girls_15_17", "pdi_girls_15_17_dis",
        "pdi_men_18_24", "pdi_men_18_24_dis",
        "pdi_women_18_24", "pdi_women_18_24_dis",
        "pdi_men_25_49", "pdi_men_25_49_dis",
        "pdi_women_25_49", "pdi_women_25_49_dis",
        "pdi_men_50_59", "pdi_men_50_59_dis",
        "pdi_women_50_59", "pdi_women_50_59_dis",
        "pdi_men_60_plus", "pdi_men_60_plus_dis",
        "pdi_women_60_plus", "pdi_women_60_plus_dis",
    ]

    ch_fields = [
        "ch_boys_0_5", "ch_boys_0_5_dis",
        "ch_girls_0_5", "ch_girls_0_5_dis",
        "ch_boys_6_14", "ch_boys_6_14_dis",
        "ch_girls_6_14", "ch_girls_6_14_dis",
        "ch_boys_15_17", "ch_boys_15_17_dis",
        "ch_girls_15_17", "ch_girls_15_17_dis",
        "ch_men_18_24", "ch_men_18_24_dis",
        "ch_women_18_24", "ch_women_18_24_dis",
        "ch_men_25_49", "ch_men_25_49_dis",
        "ch_women_25_49", "ch_women_25_49_dis",
        "ch_men_50_59", "ch_men_50_59_dis",
        "ch_women_50_59", "ch_women_50_59_dis",
        "ch_men_60_plus", "ch_men_60_plus_dis",
        "ch_women_60_plus", "ch_women_60_plus_dis",
    ]

    total_pdi = sum(getattr(obj, f, 0) or 0 for f in pdi_fields)
    total_ch = sum(getattr(obj, f, 0) or 0 for f in ch_fields)
    total_humanitarian = (obj.humanitarian_male or 0) + (obj.humanitarian_female or 0)

    set_if_exists(obj, "total_pdi", total_pdi)
    set_if_exists(obj, "total_host_community", total_ch)
    set_if_exists(obj, "total_participants", total_pdi + total_ch + total_humanitarian)


class Command(BaseCommand):
    help = "Importer les sessions EREE depuis Excel"

    def add_arguments(self, parser):
        parser.add_argument("file_path", type=str)

    def handle(self, *args, **options):
        file_path = options["file_path"]

        wb = load_workbook(file_path, data_only=True)
        ws = wb.active

        self.stdout.write(self.style.WARNING(f"Feuille utilisée : {ws.title}"))

        headers = [clean(cell.value) for cell in ws[1]]

        created = 0
        errors = 0
        skipped = 0

        for idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
            data = dict(zip(headers, row))

            if not any(row):
                continue

            try:
                region_name = clean(get_value(data, "Région"))
                cercle_name = clean(get_value(data, "Cercle"))
                commune_name = clean(get_value(data, "Commune"))

                region = find_region(region_name)
                cercle = find_cercle(cercle_name, region)
                commune = find_commune(commune_name, cercle)

                if not cercle:
                    skipped += 1
                    self.stdout.write(
                        self.style.ERROR(
                            f"Ligne {idx} ignorée : cercle introuvable → '{cercle_name}' | région='{region_name}'"
                        )
                    )
                    continue

                if not commune and commune_name:
                    commune = Commune.objects.create(
                        cercle=cercle,
                        name=commune_name,
                        code=commune_name.upper().replace(" ", "_")
                    )

                if not commune:
                    skipped += 1
                    self.stdout.write(
                        self.style.ERROR(
                            f"Ligne {idx} ignorée : commune introuvable → '{commune_name}' | cercle='{cercle_name}'"
                        )
                    )
                    continue

                session_date = clean_date(get_value(data, "Date"))

                obj = EREESession()
                obj.reference = generate_unique_eree_reference(session_date)

                set_if_exists(obj, "title", f"EREE - {commune_name}")
                set_if_exists(obj, "session_date", session_date)

                set_if_exists(obj, "longitude", clean_decimal(get_value(data, "X/Longitude/"), "lng"))
                set_if_exists(obj, "latitude", clean_decimal(get_value(data, "Y/Latitude/"), "lat"))

                set_if_exists(obj, "organisation", clean_none(get_value(data, "Organisation")))
                set_if_exists(obj, "team", clean_none(get_value(data, "Equipe")))

                set_if_exists(obj, "session_status", get_submitted_value("session_status"))
                set_if_exists(obj, "status", get_submitted_value("status"))

                set_if_exists(obj, "region", region)
                set_if_exists(obj, "cercle", cercle)
                set_if_exists(obj, "commune", commune)

                set_if_exists(obj, "region_code", getattr(region, "code", None))
                set_if_exists(obj, "cercle_code", getattr(cercle, "code", None))
                set_if_exists(obj, "commune_code", getattr(commune, "code", None))

                set_if_exists(obj, "village", clean_none(get_value(data, "Village")))
                set_if_exists(obj, "methodology", clean_none(get_value(data, "Methodologie")))
                set_if_exists(obj, "sensitization_type", clean_none(get_value(data, "Type de sensibilisation")))
                set_if_exists(obj, "civilian_subcategory", clean_none(get_value(data, "Sous-catégories de civils")))
                set_if_exists(obj, "humanitarian_org_type", clean_none(get_value(data, "Si travailleur humanitaire, preciser le type d'organisation")))
                set_if_exists(obj, "other_precision", clean_none(get_value(data, "Précision autres")))
                set_if_exists(obj, "funding_type", clean_none(get_value(data, "Source de financement (donateurs)")))

                set_if_exists(obj, "humanitarian_male", clean_int(get_value(data, "Si travailleur humanitaire, nombres de participants hommes")))
                set_if_exists(obj, "humanitarian_female", clean_int(get_value(data, "Si travailleur humanitaire, nombres de participants femmes")))

                mappings = {
                    "pdi_boys_0_5": "Garçons PDI de 0 à 5 ans",
                    "pdi_boys_0_5_dis": "Garçons PDI de 0 à 5 ans en situation d'handicap",
                    "pdi_girls_0_5": "Filles PDI de 0 à 5 ans",
                    "pdi_girls_0_5_dis": "Filles PDI de 0 à 5 ans en situation d'handicap",
                    "pdi_boys_6_14": "Garçons PDI de 6 à 14 ans",
                    "pdi_boys_6_14_dis": "Garçons PDI de 6 à 14 ans en situation d'handicap",
                    "pdi_girls_6_14": "Filles PDI de 6 à 14 ans",
                    "pdi_girls_6_14_dis": "Filles PDI de 6 à 14 ans en situation d'handicap",
                    "pdi_boys_15_17": "Garçons PDI de 15 à 17 ans",
                    "pdi_boys_15_17_dis": "Garçons PDI de 15 à 17 ans en situation d'handicap",
                    "pdi_girls_15_17": "Filles PDI de 15 à 17 ans",
                    "pdi_girls_15_17_dis": "Filles PDI de 15 à 17 ans en situation d'handicap",
                    "pdi_men_18_24": "Hommes PDI de 18 à 24 ans",
                    "pdi_men_18_24_dis": "Hommes PDI de 18 à 24 ans en situation d'handicap",
                    "pdi_women_18_24": "Femmes PDI de 18 à 24 ans",
                    "pdi_women_18_24_dis": "Femmes PDI de 18 à 24 ans en situation d'handicap",
                    "pdi_men_25_49": "Hommes PDI de 25 à 49 ans",
                    "pdi_men_25_49_dis": "Hommes PDI de 25 à 49 ans en situation d'handicap",
                    "pdi_women_25_49": "Femmes PDI de 25 à 49 ans",
                    "pdi_women_25_49_dis": "Femmes PDI de 25 à 49 ans en situation d'handicap",
                    "pdi_men_50_59": "Hommes PDI de 50 à 59 ans",
                    "pdi_men_50_59_dis": "Hommes PDI de 50 à 59 ans en situation d'handicap",
                    "pdi_women_50_59": "Femmes PDI de 50 à 59 ans",
                    "pdi_women_50_59_dis": "Femmes PDI de 50 à 59 ans en situation d'handicap",
                    "pdi_men_60_plus": "Hommes PDI de 60 ans et plus",
                    "pdi_men_60_plus_dis": "Hommes PDI de 60 et plus ans en situation d'handicap",
                    "pdi_women_60_plus": "Femmes PDI de 60 ans et plus",
                    "pdi_women_60_plus_dis": "Femmes PDI de 60 ans et plus en situation d'handicap",

                    "ch_boys_0_5": "Garçons CH de 0 à 5 ans",
                    "ch_boys_0_5_dis": "Garçons CH de 0 à 5 ans en situation d'handicap",
                    "ch_girls_0_5": "Filles CH de 0 à 5 ans",
                    "ch_girls_0_5_dis": "Filles CH de 0 à 5 ans en situation d'handicap",
                    "ch_boys_6_14": "Garçons CH de 6 à 14 ans",
                    "ch_boys_6_14_dis": "Garçons CH de 6 à 14 ans en situation d'handicap",
                    "ch_girls_6_14": "Filles CH de 6 à 14 ans",
                    "ch_girls_6_14_dis": "Filles CH de 6 à 14 ans en situation d'handicap",
                    "ch_boys_15_17": "Garçons CH de 15 à 17 ans",
                    "ch_boys_15_17_dis": "Garçons CH de 15 à 17 ans en situation d'handicap",
                    "ch_girls_15_17": "Filles CH de 15 à 17 ans",
                    "ch_girls_15_17_dis": "Filles CH de 15 à 17 ans en situation d'handicap",
                    "ch_men_18_24": "Hommes CH de 18 à 24 ans",
                    "ch_men_18_24_dis": "Hommes CH de 18 à 24 ans en situation d'handicap",
                    "ch_women_18_24": "Femmes CH de 18 à 24 ans",
                    "ch_women_18_24_dis": "Femmes CH de 18 à 24 ans en situation d'handicap",
                    "ch_men_25_49": "Hommes CH de 25 à 49 ans",
                    "ch_men_25_49_dis": "Hommes CH de 25 à 49 ans en situation d'handicap",
                    "ch_women_25_49": "Femmes CH de 25 à 49 ans",
                    "ch_women_25_49_dis": "Femmes CH de 25 à 49 ans en situation d'handicap",
                    "ch_men_50_59": "Hommes CH de 50 à 59 ans",
                    "ch_men_50_59_dis": "Hommes CH de 50 à 59 ans en situation d'handicap",
                    "ch_women_50_59": "Femmes CH de 50 à 59 ans",
                    "ch_women_50_59_dis": "Femmes CH de 50 à 59 ans en situation d'handicap",
                    "ch_men_60_plus": "Hommes CH de 60 ans et plus",
                    "ch_men_60_plus_dis": "Hommes CH de 60 et plus ans en situation d'handicap",
                    "ch_women_60_plus": "Femmes CH de 60 ans et plus",
                    "ch_women_60_plus_dis": "Femmes CH de 60 ans et plus en situation d'handicap",

                    "leaflets_adults": "Dépliants Distribués aux adultes",
                    "leaflets_children": "Dépliants Distribués aux enfants",
                }

                for model_field, excel_column in mappings.items():
                    set_if_exists(obj, model_field, clean_int(get_value(data, excel_column)))

                calculate_totals(obj)

                obj.save()
                created += 1

                self.stdout.write(self.style.SUCCESS(f"Ligne {idx} OK → {obj.reference}"))

            except Exception as e:
                errors += 1
                self.stdout.write(self.style.ERROR(f"Ligne {idx} ERREUR → {type(e).__name__}: {e}"))

        self.stdout.write(self.style.SUCCESS(f"{created} sessions EREE importées"))
        self.stdout.write(self.style.WARNING(f"{skipped} lignes ignorées"))
        self.stdout.write(self.style.WARNING(f"{errors} erreurs"))