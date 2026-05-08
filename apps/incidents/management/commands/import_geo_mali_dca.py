from openpyxl import load_workbook
from django.core.management.base import BaseCommand
from apps.geo.models import Region, Cercle, Commune


def clean(value):
    if value is None:
        return ""
    return str(value).replace("\xa0", " ").strip()


def get_value(data, possible_names):
    for name in possible_names:
        for key, value in data.items():
            if clean(key).lower() == clean(name).lower():
                return clean(value)
    return ""


class Command(BaseCommand):
    help = "Importer le découpage géographique Mali DCA"

    def add_arguments(self, parser):
        parser.add_argument("file_path", type=str)

    def handle(self, *args, **options):
        file_path = options["file_path"]

        wb = load_workbook(file_path, data_only=True)
        ws = wb.active

        headers = [clean(cell.value) for cell in ws[1]]

        self.stdout.write(self.style.WARNING("Colonnes détectées :"))
        self.stdout.write(", ".join(headers))

        regions_created = 0
        cercles_created = 0
        communes_created = 0
        skipped = 0

        for idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
            data = dict(zip(headers, row))

            region_name = get_value(data, [
                "region", "région", "Region", "Région",
                "admin1", "ADM1_FR", "NAME_1"
            ])

            cercle_name = get_value(data, [
                "cercle", "Cercle",
                "admin2", "ADM2_FR", "NAME_2"
            ])

            commune_name = get_value(data, [
                "commune", "Commune",
                "admin3", "ADM3_FR", "NAME_3"
            ])

            region_code = get_value(data, [
                "region_code", "code_region", "code région",
                "ADM1_CODE", "P_CODE_1"
            ])

            cercle_code = get_value(data, [
                "cercle_code", "code_cercle", "code cercle",
                "ADM2_CODE", "P_CODE_2"
            ])

            commune_code = get_value(data, [
                "commune_code", "code_commune", "code commune",
                "ADM3_CODE", "P_CODE_3"
            ])

            if not region_name or not cercle_name or not commune_name:
                skipped += 1
                self.stdout.write(
                    self.style.ERROR(
                        f"Ligne {idx} ignorée : region='{region_name}', cercle='{cercle_name}', commune='{commune_name}'"
                    )
                )
                continue

            region, created_region = Region.objects.get_or_create(
                name=region_name,
                defaults={"code": region_code or region_name.upper().replace(" ", "_")}
            )
            if created_region:
                regions_created += 1

            cercle, created_cercle = Cercle.objects.get_or_create(
                region=region,
                name=cercle_name,
                defaults={"code": cercle_code or cercle_name.upper().replace(" ", "_")}
            )
            if created_cercle:
                cercles_created += 1

            commune, created_commune = Commune.objects.get_or_create(
                cercle=cercle,
                name=commune_name,
                defaults={"code": commune_code or commune_name.upper().replace(" ", "_")}
            )
            if created_commune:
                communes_created += 1

        self.stdout.write(self.style.SUCCESS(f"Régions créées : {regions_created}"))
        self.stdout.write(self.style.SUCCESS(f"Cercles créés : {cercles_created}"))
        self.stdout.write(self.style.SUCCESS(f"Communes créées : {communes_created}"))
        self.stdout.write(self.style.WARNING(f"Lignes ignorées : {skipped}"))