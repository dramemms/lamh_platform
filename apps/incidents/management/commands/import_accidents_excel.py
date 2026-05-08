from openpyxl import load_workbook
from openpyxl.utils.datetime import from_excel
from django.core.management.base import BaseCommand
from django.utils.crypto import get_random_string

from apps.incidents.models import Accident
from apps.geo.models import Region, Cercle, Commune

from datetime import datetime, date, time
import re
import unicodedata


def clean(value):
    if value is None:
        return ""
    return str(value).replace("\xa0", " ").strip()


def clean_none(value):
    value = clean(value)
    return value if value else None


def normalize(value):
    value = clean(value).lower()
    value = value.replace("_cercle", "").replace("_cerle", "")
    value = unicodedata.normalize("NFKD", value)
    value = "".join(c for c in value if not unicodedata.combining(c))
    value = re.sub(r"\s+", " ", value)
    return value.strip()


def get_value(data, column_name):
    target = normalize(column_name)
    for key, value in data.items():
        if normalize(key) == target:
            return value
    return None


def clean_date(value):
    if value in [None, ""]:
        return None

    if isinstance(value, datetime):
        return value.date()

    if isinstance(value, date):
        return value

    if isinstance(value, (int, float)):
        try:
            return from_excel(value).date()
        except Exception:
            return None

    value = clean(value)

    for fmt in ("%d/%m/%Y", "%Y-%m-%d", "%d-%m-%Y"):
        try:
            return datetime.strptime(value, fmt).date()
        except Exception:
            pass

    return None


def clean_time(value):
    if value in [None, ""]:
        return None

    if isinstance(value, time):
        return value

    if isinstance(value, datetime):
        return value.time()

    value = clean(value).upper()
    value = value.replace("H", ":").replace(".", ":")

    match = re.match(r"^(\d{1,2})(?::?(\d{2}))?$", value)

    if match:
        hour = int(match.group(1))
        minute = int(match.group(2) or 0)

        if 0 <= hour <= 23 and 0 <= minute <= 59:
            return time(hour, minute)

    return None


def clean_int(value):
    if value in [None, ""]:
        return 0

    value = str(value).replace(",", ".").strip()
    match = re.search(r"\d+", value)

    if match:
        return int(match.group())

    return 0


def clean_float(value):
    if value in [None, ""]:
        return None

    try:
        return float(str(value).replace(",", "."))
    except Exception:
        return None


def generate_unique_reference(accident_date=None):
    if accident_date:
        date_part = accident_date.strftime("%Y%m%d")
    else:
        date_part = datetime.now().strftime("%Y%m%d")

    while True:
        reference = f"ACC-{date_part}-{get_random_string(8).upper()}"

        if not Accident.objects.filter(reference=reference).exists():
            return reference


def get_status_value():
    choices = [c[0] for c in Accident._meta.get_field("status").choices]

    if "SUBMITTED" in choices:
        return "SUBMITTED"

    if "SOUMIS" in choices:
        return "SOUMIS"

    return choices[0] if choices else "SUBMITTED"


def set_if_exists(obj, field_name, value):
    fields = [f.name for f in obj._meta.fields]

    if field_name in fields:
        setattr(obj, field_name, value)


def find_by_normalized_name(model, name, parent_field=None, parent_obj=None):
    name_norm = normalize(name)

    if not name_norm:
        return None

    qs = model.objects.all()

    if parent_field and parent_obj:
        qs = qs.filter(**{parent_field: parent_obj})

    for obj in qs:
        if normalize(getattr(obj, "name", "")) == name_norm:
            return obj

    return None


def find_region(name):
    return find_by_normalized_name(Region, name)


def find_cercle(name, region=None):
    cercle = find_by_normalized_name(Cercle, name, "region", region)

    if cercle:
        return cercle

    return find_by_normalized_name(Cercle, name)


def find_commune(name, cercle=None):
    commune = find_by_normalized_name(Commune, name, "cercle", cercle)

    if commune:
        return commune

    return find_by_normalized_name(Commune, name)


class Command(BaseCommand):
    help = "Importer les accidents depuis Excel"

    def add_arguments(self, parser):
        parser.add_argument("file_path", type=str)

    def handle(self, *args, **options):
        file_path = options["file_path"]

        wb = load_workbook(file_path, data_only=True)

        if "Accident form" in wb.sheetnames:
            ws = wb["Accident form"]
        else:
            ws = wb.active

        self.stdout.write(self.style.WARNING(f"Feuille utilisée : {ws.title}"))

        headers = [clean(cell.value) for cell in ws[1]]

        created = 0
        errors = 0

        for idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
            data = dict(zip(headers, row))

            if not any(row):
                continue

            try:
                region_name = clean(get_value(data, "3.2. Région"))
                cercle_name = clean(get_value(data, "3.3. Cercle"))
                commune_name = clean(get_value(data, "3.4. Commune"))

                region = find_region(region_name)
                cercle = find_cercle(cercle_name, region)
                commune = find_commune(commune_name, cercle)

                if not cercle:
                    errors += 1
                    self.stdout.write(
                        self.style.ERROR(
                            f"Ligne {idx} ignorée : cercle introuvable → '{cercle_name}' | région='{region_name}'"
                        )
                    )
                    continue

                if not commune and commune_name and cercle:
                    commune = Commune.objects.create(
                        cercle=cercle,
                        name=commune_name,
                        code=commune_name.upper().replace(" ", "_")
                    )

                if not commune:
                    errors += 1
                    self.stdout.write(
                        self.style.ERROR(
                            f"Ligne {idx} ignorée : commune introuvable → '{commune_name}' | cercle='{cercle_name}'"
                        )
                    )
                    continue

                accident_date_value = clean_date(get_value(data, "2.1. Date de l'accident"))

                accident = Accident()

                accident.reference = generate_unique_reference(accident_date_value)
                accident.status = get_status_value()

                set_if_exists(accident, "incident_id", clean_none(get_value(data, "ID de l'incident associé")))
                set_if_exists(accident, "report_date", clean_date(get_value(data, "1.2.  Date du rapport")))
                set_if_exists(accident, "organisation", clean_none(get_value(data, "1.3.Nom de l'organisation")))
                set_if_exists(accident, "reported_by", clean_none(get_value(data, "1.4. Rapporté par")))
                set_if_exists(accident, "position", clean_none(get_value(data, "1.5. Position")))
                set_if_exists(accident, "team", clean_none(get_value(data, "1.6. Equipe")))
                set_if_exists(accident, "funding_source", clean_none(get_value(data, "1.7. Source de financement")))

                set_if_exists(accident, "accident_date", accident_date_value)
                set_if_exists(accident, "accident_time", clean_time(get_value(data, "2.2. Heure de l'accident")))
                set_if_exists(accident, "zone_type", clean_none(get_value(data, "2.3. Type de zone")))
                set_if_exists(accident, "zone_type_other", clean_none(get_value(data, "2.3.1. Autre, préciser")))
                set_if_exists(accident, "number_victims", clean_int(get_value(data, "2.4. Nombre de victimes")))
                set_if_exists(accident, "other_damage", clean_none(get_value(data, "2.5. Autres dommages")))
                set_if_exists(accident, "other_damage_details", clean_none(get_value(data, "2.5.1. Si autre, préciser")))
                set_if_exists(accident, "activity_at_accident", clean_none(get_value(data, "2.6. Activité au moment de l'accident")))
                set_if_exists(accident, "activity_other", clean_none(get_value(data, "2.6.1. Si autre, préciser")))
                set_if_exists(accident, "description", clean_none(get_value(data, "2.7. Description de l' accident")))
                set_if_exists(accident, "accident_type", clean_none(get_value(data, "2.8. Type d'accident")))
                set_if_exists(accident, "device_type", clean_none(get_value(data, "2.9. Type d'engin")))
                set_if_exists(accident, "device_type_other", clean_none(get_value(data, "2.9.1. Si autre, préciser")))
                set_if_exists(accident, "device_status", clean_none(get_value(data, "2.10. Status de l'engin")))
                set_if_exists(accident, "device_marked", clean_none(get_value(data, "2.11. L'engin est-il marqué?")))

                set_if_exists(accident, "country", clean_none(get_value(data, "3.1. Pays")))
                set_if_exists(accident, "region", region)
                set_if_exists(accident, "cercle", cercle)
                set_if_exists(accident, "commune", commune)
                set_if_exists(accident, "village", clean_none(get_value(data, "3.5. Village / Quartier")))
                set_if_exists(accident, "system", clean_none(get_value(data, "3.6. Système")))
                set_if_exists(accident, "secure_access", clean_none(get_value(data, "3.7. Accès sécurisé au lieu d'accident ?")))
                set_if_exists(accident, "secure_access_reason", clean_none(get_value(data, "3.7.1. Sinon, pourquoi?")))
                set_if_exists(accident, "coordinate_source", clean_none(get_value(data, "3.8. Source de coordonnées")))
                set_if_exists(accident, "latitude", clean_float(get_value(data, "Latitude")))
                set_if_exists(accident, "longitude", clean_float(get_value(data, "Longitude")))
                set_if_exists(accident, "location_details", clean_none(get_value(data, "3.11. Détails de la localisation")))
                set_if_exists(accident, "nearest_access_point", clean_none(get_value(data, "3.12. Description du point d'accès le plus proche")))

                set_if_exists(accident, "source_last_name", clean_none(get_value(data, "4.1. Nom")))
                set_if_exists(accident, "source_first_name", clean_none(get_value(data, "4.2. Prenom")))
                set_if_exists(accident, "source_contact", clean_none(get_value(data, "4.3. Contact")))
                set_if_exists(accident, "source_gender", clean_none(get_value(data, "4.4. Sexe")))
                set_if_exists(accident, "source_age", clean_int(get_value(data, "4.5. Age")))
                set_if_exists(accident, "source_type", clean_none(get_value(data, "4.6. Type de source")))
                set_if_exists(accident, "source_other", clean_none(get_value(data, "4.7. Si autre, préciser")))

                accident.save()
                created += 1

                self.stdout.write(
                    self.style.SUCCESS(f"Ligne {idx} OK → {accident.reference}")
                )

            except Exception as e:
                errors += 1
                self.stdout.write(
                    self.style.ERROR(f"Ligne {idx} ERREUR → {type(e).__name__}: {e}")
                )

        self.stdout.write(self.style.SUCCESS(f"{created} accidents importés"))
        self.stdout.write(self.style.WARNING(f"{errors} erreurs"))