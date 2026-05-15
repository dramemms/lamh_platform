from urllib.parse import urlencode
from collections import Counter, defaultdict
import json

from apps.geo.models import Region, Cercle, Commune
from django.contrib.auth import get_user_model
from django import forms
from django.contrib import messages
from django.db.models import Count, Sum
from django.contrib.auth.decorators import login_required
from django.core.exceptions import ValidationError
from django.db.models import Q
from django.http import HttpResponse, JsonResponse
from django.shortcuts import get_object_or_404, redirect, render
from django.template.loader import render_to_string
from django.utils import timezone
from django.utils.dateparse import parse_date, parse_datetime
from django.views.decorators.csrf import csrf_exempt

from openpyxl import Workbook
from openpyxl.utils import get_column_letter
from weasyprint import HTML, CSS

from apps.core.permissions import (
    can_approve,
    can_edit_accident,
    can_program_validate,
    can_tech_validate,
    can_tech_verify,
)

from apps.eree.models import EREESession
from apps.victims.models import Victim

from .forms import AccidentEditForm
from .models import Accident, AccidentChangeLog

from apps.notifications.services import (
    notify_accident_submitted,
    notify_accident_tech_validated,
    notify_accident_program_validated,
    notify_accident_returned,
    notify_accident_approved,
)


KOBO_VICTIM_FORM_URL = "https://ee.kobotoolbox.org/x/0ZS5WfIv"


class WorkflowCommentForm(forms.Form):
    reason = forms.CharField(
        required=False,
        label="Motif",
        widget=forms.Textarea(attrs={"rows": 4, "class": "form-control"}),
    )
    comment = forms.CharField(
        required=False,
        label="Commentaire",
        widget=forms.Textarea(attrs={"rows": 4, "class": "form-control"}),
    )


def get_accident_base_queryset():
    return Accident.objects.select_related(
        "region",
        "cercle",
        "commune",
        "created_by",
        "tech_validated_by",
        "program_validated_by",
        "approved_by",
    )


def get_user_scoped_queryset(user, with_logs=False):
    queryset = get_accident_base_queryset()

    if with_logs:
        queryset = queryset.prefetch_related("change_logs", "victims")

    if user.is_superuser:
        return queryset

    if getattr(user, "region", None):
        queryset = queryset.filter(region=user.region)

    if getattr(user, "cercle", None):
        queryset = queryset.filter(cercle=user.cercle)

    if getattr(user, "commune", None):
        queryset = queryset.filter(commune=user.commune)

    return queryset.distinct()


def get_accident_or_404(user, pk, with_logs=False):
    queryset = get_user_scoped_queryset(user, with_logs=with_logs)
    return get_object_or_404(queryset, pk=pk)


def normalize_value(value):
    if value is None:
        return ""

    if hasattr(value, "pk"):
        return str(value.pk)

    if hasattr(value, "isoformat"):
        return value.isoformat()

    return str(value).strip()


def display_value(value):
    if value is None or value == "":
        return "-"

    return str(value)


def get_workflow_step_label(accident):
    status_map = {
        Accident.STATUS_SUBMITTED: "Soumission",
        Accident.STATUS_TECH_VERIFIED: "Vérification technique",
        Accident.STATUS_TECH_VALIDATED: "Validation technique",
        Accident.STATUS_PROGRAM_VALIDATED: "Validation programme",
        Accident.STATUS_RETURNED_FOR_CORRECTION: "Retourné pour correction",
        Accident.STATUS_APPROVED: "Approbation finale",
    }

    return status_map.get(accident.status, accident.get_status_display())


@login_required
def accident_list(request):
    base_queryset = get_user_scoped_queryset(request.user)
    accidents = base_queryset

    query = request.GET.get("q", "").strip()
    status = request.GET.get("status", "").strip()
    region_id = request.GET.get("region", "").strip()
    cercle_id = request.GET.get("cercle", "").strip()
    commune_id = request.GET.get("commune", "").strip()

    if query:
        accidents = accidents.filter(
            Q(reference__icontains=query)
            | Q(title__icontains=query)
            | Q(description__icontains=query)
            | Q(locality__icontains=query)
            | Q(reported_by__icontains=query)
            | Q(org_name__icontains=query)
        )

    allowed_statuses = {
        Accident.STATUS_SUBMITTED,
        Accident.STATUS_TECH_VERIFIED,
        Accident.STATUS_TECH_VALIDATED,
        Accident.STATUS_PROGRAM_VALIDATED,
        Accident.STATUS_RETURNED_FOR_CORRECTION,
        Accident.STATUS_APPROVED,
    }

    if status in allowed_statuses:
        accidents = accidents.filter(status=status)
    else:
        status = ""

    if region_id:
        accidents = accidents.filter(region_id=region_id)

    if cercle_id:
        accidents = accidents.filter(cercle_id=cercle_id)

    if commune_id:
        accidents = accidents.filter(commune_id=commune_id)

    accidents = accidents.order_by("-created_at")

    context = {
        "accidents": accidents,
        "current_status": status,
        "query": query,
        "selected_region": region_id,
        "selected_cercle": cercle_id,
        "selected_commune": commune_id,
        "count_all": base_queryset.count(),
        "count_submitted": base_queryset.filter(
            status=Accident.STATUS_SUBMITTED
        ).count(),
        "count_tech_verified": base_queryset.filter(
            status=Accident.STATUS_TECH_VERIFIED
        ).count(),
        "count_tech": base_queryset.filter(
            status=Accident.STATUS_TECH_VALIDATED
        ).count(),
        "count_program": base_queryset.filter(
            status=Accident.STATUS_PROGRAM_VALIDATED
        ).count(),
        "count_approved": base_queryset.filter(
            status=Accident.STATUS_APPROVED
        ).count(),
    }

    return render(request, "incidents/accident_list.html", context)


@login_required
def accident_detail(request, pk):
    accident = get_accident_or_404(request.user, pk, with_logs=True)

    params = {
        "d[accident_id]": accident.reference or "",
        "d[accident_date]": (
            accident.accident_date.strftime("%Y-%m-%d")
            if accident.accident_date
            else ""
        ),
    }

    kobo_victim_prefill_url = f"{KOBO_VICTIM_FORM_URL}?{urlencode(params)}"

    victims = accident.victims.all().order_by("-created_at")

    context = {
        "accident": accident,
        "victims": victims,
        "victims_count": victims.count(),
        "kobo_victim_prefill_url": kobo_victim_prefill_url,
        "can_edit_accident": can_edit_accident(request.user),
        "can_tech_verify": can_tech_verify(request.user),
        "can_tech_validate": can_tech_validate(request.user),
        "can_program_validate": can_program_validate(request.user),
        "can_approve": can_approve(request.user),
    }

    return render(request, "incidents/accident_detail.html", context)


@login_required
def accident_transition(request, pk, action):
    accident = get_accident_or_404(request.user, pk)

    try:
        if action == "submit":
            if not can_edit_accident(request.user):
                raise ValidationError("Non autorisé.")

            if accident.status != Accident.STATUS_RETURNED_FOR_CORRECTION:
                raise ValidationError("Transition invalide.")

            accident.transition_to(
                Accident.STATUS_SUBMITTED,
                user=request.user,
                comment="Accident corrigé et ressoumis.",
            )

            notify_accident_submitted(accident)

        elif action == "tech_verify":
            if not can_tech_verify(request.user):
                raise ValidationError("Non autorisé.")

            if accident.status != Accident.STATUS_SUBMITTED:
                raise ValidationError("Transition invalide.")

            accident.transition_to(
                Accident.STATUS_TECH_VERIFIED,
                user=request.user,
                comment="Vérification technique effectuée.",
            )

        elif action == "tech_validate":
            if not can_tech_validate(request.user):
                raise ValidationError("Non autorisé.")

            if accident.status != Accident.STATUS_TECH_VERIFIED:
                raise ValidationError("Transition invalide.")

            accident.transition_to(
                Accident.STATUS_TECH_VALIDATED,
                user=request.user,
            )

            notify_accident_tech_validated(accident)

        elif action == "program_validate":
            if not can_program_validate(request.user):
                raise ValidationError("Non autorisé.")

            if accident.status != Accident.STATUS_TECH_VALIDATED:
                raise ValidationError("Transition invalide.")

            accident.transition_to(
                Accident.STATUS_PROGRAM_VALIDATED,
                user=request.user,
            )

            notify_accident_program_validated(accident)

        elif action == "approve":
            if not can_approve(request.user):
                raise ValidationError("Non autorisé.")

            if accident.status != Accident.STATUS_PROGRAM_VALIDATED:
                raise ValidationError("Transition invalide.")

            accident.transition_to(
                Accident.STATUS_APPROVED,
                user=request.user,
            )

            notify_accident_approved(accident)

        else:
            raise ValidationError("Action inconnue.")

        messages.success(request, "Statut mis à jour avec succès.")

    except (ValidationError, ValueError) as e:
        messages.error(request, str(e))

    return redirect("accident_detail", pk=pk)


@login_required
def accident_reject_or_return(request, pk, action):
    accident = get_accident_or_404(request.user, pk)

    if action not in {"tech_reject", "program_reject"}:
        messages.error(request, "Action inconnue.")
        return redirect("accident_detail", pk=pk)

    form = WorkflowCommentForm(request.POST or None)

    if request.method == "POST" and form.is_valid():
        try:
            reason = (form.cleaned_data.get("reason") or "").strip()

            if not reason:
                messages.error(request, "Le motif de rejet / retour est obligatoire.")
                return render(
                    request,
                    "incidents/accident_workflow_form.html",
                    {
                        "accident": accident,
                        "form": form,
                        "action": action,
                    },
                )

            if action == "tech_reject":
                if not (
                    can_tech_verify(request.user)
                    or can_tech_validate(request.user)
                ):
                    raise ValidationError("Action non autorisée.")

                if accident.status not in [
                    Accident.STATUS_SUBMITTED,
                    Accident.STATUS_TECH_VERIFIED,
                ]:
                    raise ValidationError("Transition invalide.")

                accident.transition_to(
                    Accident.STATUS_RETURNED_FOR_CORRECTION,
                    user=request.user,
                    reason=reason,
                    comment=reason,
                )

                notify_accident_returned(accident)

                messages.success(
                    request,
                    "Le rapport a été retourné au soumissionnaire pour correction.",
                )
                return redirect("accident_detail", pk=pk)

            if action == "program_reject" and accident.status == Accident.STATUS_PROGRAM_VALIDATED:
                if not can_approve(request.user):
                    raise ValidationError("Action non autorisée.")

                accident.transition_to(
                    Accident.STATUS_TECH_VALIDATED,
                    user=request.user,
                    reason=reason,
                    comment=reason,
                )

                notify_accident_returned(accident)

                messages.success(
                    request,
                    "Le rapport a été retourné à la validation technique.",
                )
                return redirect("accident_detail", pk=pk)

            if action == "program_reject" and accident.status == Accident.STATUS_TECH_VALIDATED:
                if not can_program_validate(request.user):
                    raise ValidationError("Action non autorisée.")

                accident.transition_to(
                    Accident.STATUS_RETURNED_FOR_CORRECTION,
                    user=request.user,
                    reason=reason,
                    comment=reason,
                )

                notify_accident_returned(accident)

                messages.success(
                    request,
                    "Le rapport a été retourné au soumissionnaire pour correction.",
                )
                return redirect("accident_detail", pk=pk)

            raise ValidationError("Transition invalide.")

        except (ValidationError, ValueError) as e:
            messages.error(request, str(e))

    return render(
        request,
        "incidents/accident_workflow_form.html",
        {
            "accident": accident,
            "form": form,
            "action": action,
        },
    )


@login_required
def accident_edit(request, pk):
    accident = get_accident_or_404(request.user, pk)

    if not can_edit_accident(request.user):
        messages.error(request, "Non autorisé.")
        return redirect("accident_detail", pk=pk)

    if accident.status == Accident.STATUS_APPROVED:
        messages.error(
            request,
            "Ce rapport est approuvé et ne peut plus être modifié.",
        )
        return redirect("accident_detail", pk=pk)

    if request.method == "POST":
        form = AccidentEditForm(request.POST, instance=accident)

        if form.is_valid():
            comment = form.cleaned_data.get("comment", "").strip()

            old_accident = Accident.objects.get(pk=accident.pk)
            workflow_step_label = get_workflow_step_label(old_accident)

            old_values = {}

            for field_name in form.fields.keys():
                if field_name == "comment":
                    continue

                old_values[field_name] = getattr(old_accident, field_name, None)

            updated_accident = form.save(commit=False)
            real_changes = []

            for field_name in form.fields.keys():
                if field_name == "comment":
                    continue

                old_value = old_values.get(field_name)
                new_value = getattr(updated_accident, field_name, None)

                if normalize_value(old_value) != normalize_value(new_value):
                    field_label = form.fields[field_name].label or field_name
                    real_changes.append(
                        {
                            "field_name": field_label,
                            "old_value": display_value(old_value),
                            "new_value": display_value(new_value),
                        }
                    )

            if not real_changes:
                messages.info(request, "Aucune modification réelle détectée.")
                return redirect("accident_detail", pk=pk)

            updated_accident.save()
            form.save_m2m()

            for change in real_changes:
                AccidentChangeLog.objects.create(
                    accident=updated_accident,
                    changed_by=request.user,
                    workflow_step=workflow_step_label,
                    field_name=change["field_name"],
                    old_value=change["old_value"],
                    new_value=change["new_value"],
                    comment=comment or "-",
                )

            messages.success(request, "Rapport modifié avec succès.")
            return redirect("accident_detail", pk=pk)

        messages.error(request, "Le formulaire contient des erreurs.")

    else:
        form = AccidentEditForm(instance=accident)

    return render(
        request,
        "incidents/accident_edit.html",
        {
            "accident": accident,
            "form": form,
        },
    )


@login_required
def accident_submit(request, pk):
    return accident_transition(request, pk, "submit")


@login_required
def accident_resubmit(request, pk):
    return accident_transition(request, pk, "submit")


@login_required
def accident_tech_verify(request, pk):
    return accident_transition(request, pk, "tech_verify")


@login_required
def accident_tech_validate(request, pk):
    return accident_transition(request, pk, "tech_validate")


@login_required
def accident_program_validate(request, pk):
    return accident_transition(request, pk, "program_validate")


@login_required
def accident_approve(request, pk):
    return accident_transition(request, pk, "approve")


@login_required
def accident_tech_reject(request, pk):
    return accident_reject_or_return(request, pk, "tech_reject")


@login_required
def accident_program_reject(request, pk):
    return accident_reject_or_return(request, pk, "program_reject")


def _accident_field_verbose_name(field):
    if field.is_relation and getattr(field, "many_to_one", False):
        base = field.name[:-3] if field.name.endswith("_id") else field.name
        pretty = base.replace("_", " ").strip().capitalize()

        if base == "created_by":
            return "Créé par"

        if base == "tech_validated_by":
            return "Validé technique par"

        if base == "program_validated_by":
            return "Validé programme par"

        if base == "approved_by":
            return "Approuvé par"

        return pretty

    custom = {
        "id": "ID",
        "accident_associe_id": "ID accident associé",
        "org_name": "Organisation",
        "src_coordinates": "Coordonnées source",
        "location_gps": "GPS",
        "source_name": "Nom source",
        "source_first_name": "Prénom source",
        "source_last_name": "Nom source (famille)",
        "source_contact": "Contact source",
        "source_gender": "Genre source",
        "source_age": "Âge source",
        "source_type": "Type source",
        "reported_by": "Rapporté par",
        "submitter_first_name": "Prénom soumissionnaire",
        "submitter_last_name": "Nom soumissionnaire",
        "submitter_email": "Email soumissionnaire",
        "submitter_phone": "Téléphone soumissionnaire",
        "submitter_role": "Rôle soumissionnaire",
        "submitter_username": "Nom utilisateur soumissionnaire",
        "submitter_organization": "Organisation soumissionnaire",
        "activity_at_time": "Activité au moment de l'accident",
        "device_type": "Type d'engin",
        "device_status": "Statut de l'engin",
        "device_marked": "Engin marqué",
        "number_victims": "Nombre de victimes",
        "other_damage": "Autres dommages",
        "funding_source": "Source de financement",
        "submitted_at_kobo": "Date soumission Kobo",
        "submitted_at": "Date soumission",
        "tech_validated_at": "Date validation technique",
        "program_validated_at": "Date validation programme",
        "approved_at": "Date approbation",
        "created_at": "Créé le",
        "updated_at": "Mis à jour le",
        "validation_comment": "Commentaire validation",
        "correction_comment": "Commentaire correction",
        "rejection_reason": "Motif rejet/retour",
        "secure_access": "Accès sécurisé",
        "category": "Catégorie",
        "status": "Statut",
        "title": "Titre",
        "description": "Description",
        "reference": "Référence",
        "accident_date": "Date accident",
        "accident_time": "Heure accident",
        "latitude": "Latitude",
        "longitude": "Longitude",
        "locality": "Localité",
        "team": "Équipe",
    }

    if field.name in custom:
        return custom[field.name]

    return (field.verbose_name or field.name).replace("_", " ").capitalize()


def _accident_cell_value(obj, field):
    value = getattr(obj, field.name)

    if value is None:
        return ""

    if field.name == "status":
        try:
            return obj.get_status_display()
        except Exception:
            return str(value)

    if field.is_relation and getattr(field, "many_to_one", False):
        return str(value) if value else ""

    if isinstance(value, bool):
        return "Oui" if value else "Non"

    if hasattr(value, "strftime"):
        try:
            return value.strftime("%d/%m/%Y %H:%M:%S")
        except Exception:
            return str(value)

    return str(value)


@login_required
def export_accidents_excel(request):
    accidents = get_user_scoped_queryset(request.user).filter(
        status=Accident.STATUS_APPROVED
    )

    query = request.GET.get("q", "").strip()
    status = request.GET.get("status", "").strip()
    region_id = request.GET.get("region", "").strip()
    cercle_id = request.GET.get("cercle", "").strip()
    commune_id = request.GET.get("commune", "").strip()

    if query:
        accidents = accidents.filter(
            Q(reference__icontains=query)
            | Q(title__icontains=query)
            | Q(description__icontains=query)
            | Q(locality__icontains=query)
            | Q(reported_by__icontains=query)
            | Q(org_name__icontains=query)
        )

    allowed_statuses = {
        Accident.STATUS_SUBMITTED,
        Accident.STATUS_TECH_VERIFIED,
        Accident.STATUS_TECH_VALIDATED,
        Accident.STATUS_PROGRAM_VALIDATED,
        Accident.STATUS_RETURNED_FOR_CORRECTION,
        Accident.STATUS_APPROVED,
    }

    if status in allowed_statuses:
        accidents = accidents.filter(status=status)

    if region_id:
        accidents = accidents.filter(region_id=region_id)

    if cercle_id:
        accidents = accidents.filter(cercle_id=cercle_id)

    if commune_id:
        accidents = accidents.filter(commune_id=commune_id)

    accidents = accidents.order_by("-created_at")

    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Accidents"

    excluded_fields = [
        "raw_payload",
        "kobo_uuid",
        "kobo_asset_uid",
        "kobo_submission_id",
        "is_synced",
        "synced_at",
    ]

    fields = [
        field for field in Accident._meta.fields
        if field.name not in excluded_fields
    ]

    headers = [_accident_field_verbose_name(field) for field in fields]
    sheet.append(headers)

    for obj in accidents:
        row = [_accident_cell_value(obj, field) for field in fields]
        sheet.append(row)

    for col in range(1, sheet.max_column + 1):
        max_length = 0
        col_letter = get_column_letter(col)

        for row in range(1, sheet.max_row + 1):
            value = sheet.cell(row=row, column=col).value

            if value:
                max_length = max(max_length, len(str(value)))

        sheet.column_dimensions[col_letter].width = min(max_length + 2, 60)

    response = HttpResponse(
        content_type=(
            "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )
    )

    response["Content-Disposition"] = 'attachment; filename="accidents.xlsx"'

    workbook.save(response)

    return response


def accident_dashboard(request):
    accidents = get_user_scoped_queryset(request.user).filter(
        status=Accident.STATUS_APPROVED
    )

    organisation = request.GET.get("organisation", "").strip()
    accident_type = request.GET.get("accident_type", "").strip()
    region = request.GET.get("region", "").strip()
    cercle = request.GET.get("cercle", "").strip()
    commune = request.GET.get("commune", "").strip()
    periode = request.GET.get("periode", "").strip()

    if organisation:
        accidents = accidents.filter(org_name__iexact=organisation)

    if accident_type:
        accidents = accidents.filter(category__iexact=accident_type)

    if region:
        accidents = accidents.filter(region__name__iexact=region)

    if cercle:
        accidents = accidents.filter(cercle__name__iexact=cercle)

    if commune:
        accidents = accidents.filter(commune__name__iexact=commune)

    if periode and periode.isdigit():
        accidents = accidents.filter(accident_date__year=int(periode))

    accidents = accidents.order_by("-created_at")
    total_accidents = accidents.count()

    filter_queryset = get_user_scoped_queryset(request.user)

    organisations = (
        filter_queryset.exclude(org_name__isnull=True)
        .exclude(org_name__exact="")
        .values_list("org_name", flat=True)
        .distinct()
        .order_by("org_name")
    )

    accident_types = (
        filter_queryset.exclude(category__isnull=True)
        .exclude(category__exact="")
        .values_list("category", flat=True)
        .distinct()
        .order_by("category")
    )

    regions = (
        filter_queryset.filter(region__isnull=False)
        .values_list("region__name", flat=True)
        .distinct()
        .order_by("region__name")
    )

    cercles = (
        filter_queryset.filter(cercle__isnull=False)
        .values_list("cercle__name", flat=True)
        .distinct()
        .order_by("cercle__name")
    )

    communes = (
        filter_queryset.filter(commune__isnull=False)
        .values_list("commune__name", flat=True)
        .distinct()
        .order_by("commune__name")
    )

    def find_in_payload(payload, possible_keys):
        if not payload or not isinstance(payload, dict):
            return None

        for key in possible_keys:
            if key in payload and payload.get(key):
                return payload.get(key)

        for value in payload.values():
            if isinstance(value, dict):
                found = find_in_payload(value, possible_keys)

                if found:
                    return found

        return None

    month_order = [
        "janvier",
        "février",
        "mars",
        "avril",
        "mai",
        "juin",
        "juillet",
        "août",
        "septembre",
        "octobre",
        "novembre",
        "décembre",
    ]

    month_map = {
        1: "janvier",
        2: "février",
        3: "mars",
        4: "avril",
        5: "mai",
        6: "juin",
        7: "juillet",
        8: "août",
        9: "septembre",
        10: "octobre",
        11: "novembre",
        12: "décembre",
    }

    month_counter = defaultdict(int)

    for accident in accidents:
        if accident.accident_date:
            month_counter[month_map[accident.accident_date.month]] += 1

    month_labels = month_order
    month_values = [month_counter.get(month, 0) for month in month_order]

    type_counter = Counter()

    for accident in accidents:
        label = getattr(accident, "category", None)

        if not label:
            label = find_in_payload(
                getattr(accident, "raw_payload", None),
                [
                    "category",
                    "accident_type",
                    "type_accident",
                    "g_accident/category",
                    "g_accident/accident_type",
                    "g_accident/type_accident",
                ],
            )

        if not label:
            label = "(Non renseigné)"

        type_counter[str(label)] += 1

    accident_type_labels = list(type_counter.keys())
    accident_type_values = list(type_counter.values())

    engine_counter = Counter()

    for accident in accidents:
        label = getattr(accident, "device_type", None)

        if not label:
            label = find_in_payload(
                getattr(accident, "raw_payload", None),
                [
                    "device_type",
                    "type_engin",
                    "engin_type",
                    "device",
                    "g_accident/device_type",
                    "g_accident/type_engin",
                    "g_device/device_type",
                    "g_device/type_engin",
                    "g_device/engin_type",
                ],
            )

        if not label:
            label = "(Non renseigné)"

        engine_counter[str(label)] += 1

    engine_table = []

    for label, count in engine_counter.most_common():
        percent = round((count / total_accidents) * 100, 2) if total_accidents else 0

        engine_table.append(
            {
                "label": label,
                "count": count,
                "percent": str(percent).replace(".", ","),
            }
        )

    cercle_counter = Counter()

    for accident in accidents:
        label = accident.cercle.name if accident.cercle else "(Sans cercle)"
        cercle_counter[label] += 1

    cercle_labels = list(cercle_counter.keys())
    cercle_values = list(cercle_counter.values())

    CERCLE_COORDS = {
        "Bamako": [12.6392, -8.0029],
        "Bandiagara": [14.3500, -3.6100],
        "Douentza": [15.0016, -2.9498],
        "Mopti": [14.4843, -4.1820],
        "Sévaré": [14.5274, -4.0970],
        "Koro": [14.0631, -3.0787],
        "Bankass": [14.0717, -3.5144],
        "Djenné": [13.9061, -4.5533],
        "Ténenkou": [14.4572, -4.9169],
        "Youwarou": [15.3689, -4.2622],
        "Gao": [16.2717, -0.0447],
        "Ansongo": [15.6597, 0.5022],
        "Bourem": [16.9541, -0.3486],
        "Ménaka": [15.9182, 2.4022],
        "Tombouctou": [16.7666, -3.0026],
        "Goundam": [16.4145, -3.6708],
        "Diré": [16.2570, -3.4010],
        "Niafunké": [15.9322, -3.9906],
        "Kayes": [14.4469, -11.4445],
        "Koulikoro": [12.8627, -7.5599],
        "Ségou": [13.4317, -6.2157],
        "Sikasso": [11.3176, -5.6665],
        "Boré": [14.1850, -3.9220],
        "Boni": [15.0700, -2.2200],
        "Aguel-Hoc": [19.4667, 1.4167],
        "Bougouni": [11.4167, -7.4833],
        "Baraouéli": [14.3111, -6.0403],
        "Bla": [12.9442, -5.7561],
        "Macina": [13.9642, -5.3578],
        "Niono": [14.2500, -5.9833],
        "San": [13.3033, -4.8956],
        "Tominian": [13.2878, -3.6767],
        "Almoustarat": [17.0500, -0.1500],
        "Gourma-Rharous": [16.0833, -1.7667],
    }

    map_points = []

    for accident in accidents:
        lat = None
        lng = None

        if accident.latitude is not None and accident.longitude is not None:
            lat = float(accident.latitude)
            lng = float(accident.longitude)

        elif (
            accident.commune
            and accident.commune.latitude is not None
            and accident.commune.longitude is not None
        ):
            lat = float(accident.commune.latitude)
            lng = float(accident.commune.longitude)

        elif (
            accident.cercle
            and accident.cercle.latitude is not None
            and accident.cercle.longitude is not None
        ):
            lat = float(accident.cercle.latitude)
            lng = float(accident.cercle.longitude)

        elif (
            accident.region
            and accident.region.latitude is not None
            and accident.region.longitude is not None
        ):
            lat = float(accident.region.latitude)
            lng = float(accident.region.longitude)

        elif accident.cercle and accident.cercle.name in CERCLE_COORDS:
            lat, lng = CERCLE_COORDS[accident.cercle.name]

        if lat is not None and lng is not None:
            map_points.append(
                {
                    "reference": accident.reference,
                    "cercle": accident.cercle.name if accident.cercle else "Sans cercle",
                    "commune": accident.commune.name if accident.commune else "",
                    "region": accident.region.name if accident.region else "",
                    "count": 1,
                    "lat": lat,
                    "lng": lng,
                }
            )

    context = {
        "total_accidents": total_accidents,
        "organisations": organisations,
        "accident_types": accident_types,
        "regions": regions,
        "cercles": cercles,
        "communes": communes,
        "selected_organisation": organisation,
        "selected_accident_type": accident_type,
        "selected_region": region,
        "selected_cercle": cercle,
        "selected_commune": commune,
        "selected_periode": periode,
        "month_labels_json": json.dumps(month_labels),
        "month_values_json": json.dumps(month_values),
        "accident_type_labels_json": json.dumps(accident_type_labels),
        "accident_type_values_json": json.dumps(accident_type_values),
        "engine_table": engine_table,
        "cercle_labels_json": json.dumps(cercle_labels),
        "cercle_values_json": json.dumps(cercle_values),
        "map_points_json": json.dumps(map_points),
    }

    return render(request, "incidents/accident_dashboard.html", context)


def rate(part, total):
    if not total:
        return 0

    return round((part / total) * 100, 1)


@login_required
def export_accident_dashboard_pdf(request):
    accidents_count = Accident.objects.count()
    victims_count = Victim.objects.count()
    eree_count = EREESession.objects.count()

    accidents_approved = Accident.objects.filter(status="APPROVED").count()
    victims_approved = Victim.objects.filter(status="APPROVED").count()
    eree_approved = EREESession.objects.filter(status="APPROVED").count()

    accidents_pending = accidents_count - accidents_approved
    victims_pending = victims_count - victims_approved
    eree_pending = eree_count - eree_approved

    context = {
        "generated_at": timezone.now(),
        "total_records": accidents_count + victims_count + eree_count,
        "accidents_count": accidents_count,
        "accidents_approved": accidents_approved,
        "accidents_pending": accidents_pending,
        "accidents_approval_rate": rate(accidents_approved, accidents_count),
        "victims_count": victims_count,
        "victims_approved": victims_approved,
        "victims_pending": victims_pending,
        "victims_approval_rate": rate(victims_approved, victims_count),
        "eree_count": eree_count,
        "eree_approved": eree_approved,
        "eree_pending": eree_pending,
        "accidents_by_status": (
            Accident.objects.values("status")
            .annotate(total=Count("id"))
            .order_by("-total")
        ),
        "accidents_by_region": (
            Accident.objects.values("region__name")
            .annotate(total=Count("id"))
            .order_by("-total")[:10]
        ),
        "victims_by_status": (
            Victim.objects.values("status")
            .annotate(total=Count("id"))
            .order_by("-total")
        ),
        "victims_by_gender": (
            Victim.objects.values("victim_sex")
            .annotate(total=Count("id"))
            .order_by("-total")
        ),
        "victims_by_type": (
            Victim.objects.values("victim_type")
            .annotate(total=Count("id"))
            .order_by("-total")
        ),
        "eree_by_status": (
            EREESession.objects.values("status")
            .annotate(total=Count("id"))
            .order_by("-total")
        ),
        "eree_by_organisation": (
            EREESession.objects.values("organisation")
            .annotate(total=Count("id"))
            .order_by("-total")[:10]
        ),
        "eree_by_region": (
            EREESession.objects.values("region__name")
            .annotate(total=Count("id"))
            .order_by("-total")[:10]
        ),
        "eree_participants_total": (
            EREESession.objects.aggregate(total=Sum("total_participants"))["total"] or 0
        ),
    }

    html_string = render_to_string(
        "incidents/dashboard_pdf.html",
        context,
        request=request,
    )

    response = HttpResponse(content_type="application/pdf")
    response["Content-Disposition"] = 'attachment; filename="rapport_dashboard_lamh_complet.pdf"'

    HTML(string=html_string, base_url=request.build_absolute_uri()).write_pdf(
        response,
        stylesheets=[
            CSS(
                string="""
                @page {
                    size: A4 landscape;
                    margin: 1.2cm;
                }
                """
            )
        ],
    )

    return response


def get_kobo_value(data, *keys):
    for key in keys:
        value = data.get(key)

        if value not in [None, ""]:
            return value

    return None


def parse_kobo_date(value):
    if not value:
        return None

    dt = parse_datetime(value)

    if dt:
        return dt.date()

    d = parse_date(value)

    if d:
        return d

    return None


@csrf_exempt
def kobo_accident_webhook(request):
    if request.method != "POST":
        return JsonResponse({"error": "POST only"}, status=405)

    try:
        data = json.loads(request.body)

        print("===== KOBO ACCIDENT DATA =====")
        print(data)
        print("===== KOBO KEYS =====")
        print(list(data.keys()))

        accident_date_value = get_kobo_value(
            data,
            "accident_details/accident_date",
            "accident_date",
            "date_accident",
            "date",
        )

        accident_date = parse_kobo_date(accident_date_value)

        if not accident_date:
            return JsonResponse(
                {
                    "error": "Champ obligatoire manquant : accident_date",
                    "received_keys": list(data.keys()),
                },
                status=400,
            )

        kobo_id = (
            data.get("_id")
            or data.get("_uuid")
            or data.get("meta/instanceID")
        )

        reference = get_kobo_value(
            data,
            "reporting/accident_id",
            "accident_id",
            "reference",
        ) or f"ACC-{timezone.now().strftime('%Y%m%d')}-{str(kobo_id)[-6:]}"

        org_name = get_kobo_value(
            data,
            "reporting/org_name",
            "reporting/organisation",
            "reporting/organization",
            "org_name",
            "organisation",
            "organization",
        ) or "Organisation"

        title = f"{reference} - {org_name}"

        region_code = str(
            get_kobo_value(data, "location/region", "region") or ""
        ).strip()

        cercle_code = str(
            get_kobo_value(data, "location/cercle", "cercle") or ""
        ).strip()

        commune_code = str(
            get_kobo_value(data, "location/commune", "commune") or ""
        ).strip()

        region_obj = None
        cercle_obj = None
        commune_obj = None

        if region_code:
            region_obj = (
                Region.objects.filter(code=region_code).first()
                or Region.objects.filter(name__iexact=region_code).first()
            )

        if cercle_code:
            cercle_obj = (
                Cercle.objects.filter(code=cercle_code).first()
                or Cercle.objects.filter(name__iexact=cercle_code).first()
            )

        if commune_code:
            commune_obj = (
                Commune.objects.filter(code=commune_code).first()
                or Commune.objects.filter(name__iexact=commune_code).first()
            )

        if commune_obj:
            if not cercle_obj:
                cercle_obj = commune_obj.cercle

            if not region_obj and commune_obj.cercle:
                region_obj = commune_obj.cercle.region

        if cercle_obj and not region_obj:
            region_obj = cercle_obj.region

        if not region_obj or not cercle_obj:
            return JsonResponse(
                {
                    "error": "Géographie introuvable ou incomplète",
                    "region_recue": region_code,
                    "cercle_recu": cercle_code,
                    "commune_recue": commune_code,
                    "region_trouvee": str(region_obj) if region_obj else None,
                    "cercle_trouve": str(cercle_obj) if cercle_obj else None,
                    "commune_trouvee": str(commune_obj) if commune_obj else None,
                },
                status=400,
            )

        latitude = get_kobo_value(
            data,
            "location/latitude",
            "latitude",
        )

        longitude = get_kobo_value(
            data,
            "location/longitude",
            "longitude",
        )

        gps = get_kobo_value(
            data,
            "location/location_gps",
            "location/gps",
            "location_gps",
            "gps",
        )

        if gps and (not latitude or not longitude):
            try:
                parts = str(gps).split()
                latitude = parts[0]
                longitude = parts[1]
            except Exception:
                pass

        category = get_kobo_value(
            data,
            "accident_details/type_accident",
            "type_accident",
            "category",
            "accident_type",
        ) or "Autre"

        submitted_by = get_kobo_value(
            data,
            "_submitted_by",
            "submitter_username",
        )

        created_by_user = None
        submitter_first_name = None
        submitter_last_name = None
        submitter_email = None
        submitter_phone = None
        submitter_organization = None
        submitter_role = None

        if submitted_by:
            User = get_user_model()

            created_by_user = (
                User.objects.filter(username__iexact=submitted_by).first()
                or User.objects.filter(email__iexact=submitted_by).first()
            )

            if created_by_user:
                submitter_first_name = created_by_user.first_name
                submitter_last_name = created_by_user.last_name
                submitter_email = created_by_user.email
                submitter_phone = getattr(created_by_user, "phone", None)
                submitter_organization = getattr(
                    created_by_user,
                    "organization",
                    None,
                )
                submitter_role = getattr(created_by_user, "role", None)

        accident, created = Accident.objects.update_or_create(
            kobo_submission_id=str(kobo_id),
            defaults={
                "reference": reference,
                "title": title,
                "accident_date": accident_date,
                "accident_time": get_kobo_value(
                    data,
                    "accident_details/accident_time",
                    "accident_time",
                ),
                "category": category,
                "number_victims": get_kobo_value(
                    data,
                    "accident_details/number_victims",
                    "number_victims",
                ),
                "other_damage": get_kobo_value(
                    data,
                    "accident_details/other_damage",
                    "other_damage",
                ),
                "activity_at_time": get_kobo_value(
                    data,
                    "accident_details/activity_at_time",
                    "activity_at_time",
                ),
                "description": get_kobo_value(
                    data,
                    "accident_details/description",
                    "description",
                ),
                "device_type": get_kobo_value(
                    data,
                    "accident_details/device_type",
                    "device_type",
                ),
                "device_status": get_kobo_value(
                    data,
                    "accident_details/device_status",
                    "device_status",
                ),
                "device_marked": get_kobo_value(
                    data,
                    "accident_details/device_marked",
                    "device_marked",
                ),
                "report_date": parse_kobo_date(
                    get_kobo_value(
                        data,
                        "reporting/report_date",
                        "reporting/date_report",
                        "reporting/date_rapport",
                        "report_date",
                        "date_report",
                        "date_rapport",
                    )
                ),
                "org_name": org_name,
                "reported_by": get_kobo_value(
    data,
    "reporting/reported_by",
    "reported_by",
),
"position": get_kobo_value(
    data,
    "reporting/position",
    "position",
),
                "team": get_kobo_value(
    data,
    "reporting/team_001",
    "team_001",
    "reporting/team",
    "team",
),
                "funding_source": get_kobo_value(
    data,
    "reporting/funding_source",
    "funding_source",
),
                "country": get_kobo_value(
                    data,
                    "location/country",
                    "location/pays",
                    "country",
                    "pays",
                ),
                "region": region_obj,
                "cercle": cercle_obj,
                "commune": commune_obj,
                "locality": get_kobo_value(
                    data,
                    "location/locality",
                    "location/village",
                    "location/village_quartier",
                    "locality",
                    "village",
                ),
                "latitude": latitude,
                "longitude": longitude,
                "secure_access": get_kobo_value(
                    data,
                    "location/secure_access",
                    "secure_access",
                ),
                "source_first_name": get_kobo_value(
                    data,
                    "source_details/source_first_name",
                    "source_first_name",
                ),
                "source_last_name": get_kobo_value(
                    data,
                    "source_details/source_last_name",
                    "source_last_name",
                ),
                "source_contact": get_kobo_value(
                    data,
                    "source_details/source_contact",
                    "source_contact",
                ),
                "source_gender": get_kobo_value(
                    data,
                    "source_details/source_gender",
                    "source_gender",
                ),
                "source_age": get_kobo_value(
                    data,
                    "source_details/source_age",
                    "source_age",
                ),
                "source_type": get_kobo_value(
                    data,
                    "source_details/source_type",
                    "source_type",
                ),
                "submitter_username": submitted_by,
                "submitter_first_name": submitter_first_name,
                "submitter_last_name": submitter_last_name,
                "submitter_email": get_kobo_value(
    data,
    "reporting/Adresse_Email",
    "Adresse_Email",
    "adresse_email",
    "email",
    "reporting/email",
) or submitter_email,
                "submitter_phone": submitter_phone,
                "submitter_organization": submitter_organization,
                "submitter_role": submitter_role,
                "created_by": created_by_user,
                "source": Accident.SOURCE_KOBO,
                "raw_payload": data,
                "status": Accident.STATUS_SUBMITTED,
                "submitted_at_kobo": (
                    parse_datetime(data.get("end"))
                    if data.get("end")
                    else None
                ),
            },
        )

        if not accident.status:
            accident.status = Accident.STATUS_SUBMITTED
            accident.save(update_fields=["status"])

        if created:
            notify_accident_submitted(accident)

        return JsonResponse(
            {
                "status": "success",
                "created": created,
                "accident_id": accident.id,
                "reference": accident.reference,
            },
            status=201,
        )

    except Exception as e:
        print("WEBHOOK ERROR:", str(e))
        return JsonResponse({"error": str(e)}, status=500)